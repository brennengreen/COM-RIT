from openpyxl import Workbook, load_workbook


# Load the workbook that this will save too and create a variable for the opentracker sheet
target_filename = "consolidated.xlsx"
target_book = load_workbook(filename=target_filename)
target_sheet = target_book["opentracker"]

# Load the dataset in which we're attempting to filter
opentracker_book = load_workbook(filename="opentrack_raw.xlsx")
datasheet = opentracker_book.active

# Create a list of all ip addressed that aren't from test stations
ip_list = {index+1:cell.value for index,cell in enumerate(datasheet["B:B"])}

def copy_row(sheet, copy_row, row_num):
    """Copy one row into another sheet 

    :param sheet sheet: The sheet in which to copy the row to
    :param row copy_row: The row value to copy
    :param int row_num: The row number to insert the row at

    """

    current_column = 1
    for cell in copy_row:
        new_cell = sheet.cell(row=row_num, column=current_column, value=cell.value)
        current_column += 1

current_row = 2
for row in ip_list:
    if ip_list[row] is None or "128.163" in ip_list[row]:
        continue
    else:
        print(datasheet[row])
        copy_row(target_sheet, datasheet[row], current_row)
        current_row += 1


target_book.save(filename=target_filename)