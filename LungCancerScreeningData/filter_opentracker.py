from openpyxl import Workbook, load_workbook

# Load the consolidated workbook and set up variables for the target and datasheets
target_filename = "consolidated.xlsx"
target_book = load_workbook(filename=target_filename)
target_sheet = target_book["opentracker_unique"]

mouseflow_sheet = target_book["mouseflow_unique"]
opentracker_sheet = target_book["opentracker"]


# Create two lists for all the opentrack ip addresses and the unique mouseflow addresses
mouseflow_ip_list = [cell.value for cell in mouseflow_sheet["L:L"]]
opentracker_ip_list = [(index+1,cell.value) for index,cell in enumerate(opentracker_sheet["B:B"])]

def copy_row(sheet, copy_row, row_num):
    """Copy one row into another sheet 

    :param sheet sheet: The sheet in which to copy the row to
    :param row copy_row: The row value to copy
    :param int row_num: The row number to insert the row at

    """

    current_column = 1
    for cell in copy_row:
        new_cell = sheet.cell(row=row_num, column=current_column, value=cell.value)
        current_column += 1

# Copy all rows that contain and ip address that is not in the mouseflow datasheet
current_row = 1
for row,address in opentracker_ip_list:
    if address not in mouseflow_ip_list:
        copy_row(target_sheet, opentracker_sheet[row], current_row)
        current_row += 1

# Save the targetbook
target_book.save(filename=target_filename)