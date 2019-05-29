from openpyxl import Workbook, load_workbook

# Load the consolidated workbook and initialize variable for each respective sheet
target_book = load_workbook(filename="consolidated.xlsx")
mouseflow_sheet = target_book["mouseflow"]
opentracker_sheet = target_book["opentracker"]

# Load each respective workbooks
mouseflow_book = load_workbook(filename="mouseflow_raw.xlsx")
mouseflow_datasheet = mouseflow_book.active
opentracker_book = load_workbook(filename="opentrack_raw.xlsx")
opentracker_datasheet = opentracker_book.active

# Create two list of IP Addressed from each sheet to compare later
mouseflow_ip_list = [cell.value for cell in mouseflow_datasheet["L:L"]]
opentracker_ip_list = [cell.value for cell in opentracker_datasheet["B:B"]]

def copy_row(sheet, copy_row, row_num):
    """Copy one row into another sheet 

    :param sheet sheet: The sheet in which to copy the row to
    :param row copy_row: The row value to copy
    :param int row_num: The row number to insert the row at

    """

    current_column = 1
    for cell in copy_row:
        new_cell = sheet.cell(row=row_num, column=current_column, value=cell.value)
        current_column += 1

# Iterate through opentracker_ip_list and copy each row that has an IP address that is in the mouseflow data
current_row = 2
for index,ip_addr in enumerate(opentracker_ip_list):
    if ip_addr in mouseflow_ip_list and "128.163" not in ip_addr:
        row = index + 1
        copy_row(opentracker_sheet, opentracker_datasheet[row], current_row)
        current_row += 1
        #print("Row:", row, "Address:", ip_addr)




# Save data to consolidated workbook
target_book.save(filename = "consolidated.xlsx")
